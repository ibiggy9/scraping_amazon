import openpyxl

def american():
    #path to sales report from vendor central
    wb = openpyxl.load_workbook('path to report')
    ws = wb['US_Detail View_Sales Diagnostic']
    mylist = []
    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:    
            mylist.append(cell.value)

    US_ASINS = []
    condition = 'B0'
    for i in mylist:
        if condition in i:
            US_ASINS.append(i)
    url_default = 'https://www.amazon.com/dp/'
    url_list = []
    for i in US_ASINS:
        url_list.append(f'{url_default}{i}')
    
    return url_list

def canadian():
    wb = openpyxl.load_workbook('Path to report')
    ws = wb['CA_Detail View_Sales Diagnostic']
    mylist = []
    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:    
            mylist.append(cell.value)

    CA_ASINS = []
    condition = 'B0'
    for i in mylist:
        if condition in i:
            CA_ASINS.append(i)
    url_default = 'https://www.amazon.ca/dp/'
    url_list = []
    for i in CA_ASINS:
        url_list.append(f'{url_default}{i}')
  
    
    return url_list
  
    



